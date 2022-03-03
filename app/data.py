from pathlib import Path
from download import Download
FILEURL = "https://www.city.kofu.yamanashi.jp/joho/opendata/shisetsu/documents/shiteibunkazai_20200306.xlsx"
import pandas as pd
from openpyxl import load_workbook
import neologdn
from datetime import datetime, timedelta
import re
from jeraconv import jeraconv
j2w = jeraconv.J2W()

def format_date(st):
    if isinstance(st, int):
        if st > 59:
            st -= 1
        return pd.Timestamp("1899-12-31") + pd.Timedelta(st, unit="d")
    elif isinstance(st[0], str):
        _q = re.search("(平成元年)(\d+月\d+日)",st)
        nen = j2w.convert(_q[1])
        gappi = _q[2]
        return datetime.strptime(str(nen) + "年" + gappi, "%Y年%m月%d日")
    else:
        return datetime.strptime(st, "%Y年%m月%d日")

class Kofu_bunkazai():
    BASE_DIR = Path(__file__).absolute().parent.parent
    DATA_DIR = BASE_DIR / "data"

    def __init__(self):
        if not self.DATA_DIR.exists():
            self.DATA_DIR.mkdir()
        
        d = Download(FILEURL, self.DATA_DIR)
        d.download()
        self.fname = self.DATA_DIR / d.name
    
    def create_df(self):
        df = pd.read_excel(self.fname, sheet_name="文化財一覧", header=2,usecols="B:G")
        df = df.dropna()
        df.columns = [neologdn.normalize(i) for i in df.columns]
        df["指定年月日"] = df["指定年月日"].map(format_date)
        self.df = df
    
    def get_version(self):
        wb = load_workbook(self.fname)
        sh = wb.worksheets[0]
        return sh.cell(204,4).value

    def query(self, keywords):
        return self.df.loc[self.df["所在地"].str.contains(keywords) | self.df["名称"].str.contains(keywords)]